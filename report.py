import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE = r""
OUTPUT_FILE = r"" 

weeks_config = [
    {"week": "Week 1", "milestone": "Foundation", "start": "2025-10-27", "end": "2025-11-02"},
    {"week": "Week 2", "milestone": "System Design", "start": "2025-11-03", "end": "2025-11-09"},
    {"week": "Week 3", "milestone": "Application Development", "start": "2025-11-10", "end": "2025-11-16"},
    {"week": "Week 4-5", "milestone": "Testing", "start": "2025-11-17", "end": "2025-11-30"},
    {"week": "Week 6", "milestone": "Documentation", "start": "2025-12-01", "end": "2025-12-05"},
]

df = pd.read_csv(INPUT_FILE, sep="\t")

def parse_date(d):
    try:
        return datetime.strptime(d, "%b %d, %Y")
    except Exception:
        return None

df["Start date"] = df["Start date"].apply(parse_date)
df["Target date"] = df["Target date"].apply(parse_date)

headers = ["Task"]
for w in weeks_config:
    start = datetime.strptime(w["start"], "%Y-%m-%d")
    end = datetime.strptime(w["end"], "%Y-%m-%d")
    date_cols = [(start + timedelta(days=i)).strftime("%m/%d") for i in range((end - start).days + 1)]
    headers += date_cols + [w["week"], w["milestone"], "Status", "Assignees", "Size", "Estimate", "Priority", "Finished"]

wb = Workbook()
ws = wb.active
ws.title = "Weekly Roadmap"
ws.append(headers)

orange_fill = PatternFill(start_color="FFF06A2E", end_color="FFF06A2E", fill_type="solid")
blue_fill = PatternFill(start_color="FF8EDAFA", end_color="FF8EDAFA", fill_type="solid")
black_fill = PatternFill(start_color="FFD84D35", end_color="FFD84D35", fill_type="solid")

for _, r in df.iterrows():
    task = r["Title"]
    start = r["Start date"]
    end = r["Target date"]
    status = str(r["Status"]).lower()
    milestone = str(r["Milestone"]).strip()
    assignees = str(r["Assignees"])
    size = r["Size"]
    estimate = r["Estimate"]
    priority = r["Priority"]

    row_values = [task]

    for w in weeks_config:
        start_w = datetime.strptime(w["start"], "%Y-%m-%d")
        end_w = datetime.strptime(w["end"], "%Y-%m-%d")
        days = [(start_w + timedelta(days=i)) for i in range((end_w - start_w).days + 1)]

        if milestone != w["milestone"]:
            for _ in days:
                row_values.append("")
            row_values += ["", "", "", "", "", "", "", "FINISHED"]
            continue

        for d in days:
            if not start:
                row_values.append("")
                continue

            if "done" in status:
                if start <= d < end:
                    row_values.append("orange")
                elif d == end:
                    row_values.append("blue")
                else:
                    row_values.append("")
            else:
                row_values.append("orange" if d >= start else "")

        row_values += [w["week"], w["milestone"], status, assignees, size, estimate, priority, "FINISHED"]

    ws.append(row_values)
    row_idx = ws.max_row

    for col_idx, val in enumerate(row_values, start=1):
        cell = ws.cell(row_idx, col_idx)
        if val == "orange":
            cell.value = None
            cell.fill = orange_fill
        elif val == "blue":
            cell.value = None
            cell.fill = blue_fill
        elif val == "FINISHED":
            cell.value = None
            cell.fill = black_fill
        else:
            cell.value = val

for column_cells in ws.columns:
    column_letter = get_column_letter(column_cells[0].column)
    max_length = 0
    for cell in column_cells:
        try:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

df["Estimate"] = pd.to_numeric(df["Estimate"], errors="coerce").fillna(0)
df["Assignees"] = df["Assignees"].fillna("")
df["Milestone"] = df["Milestone"].fillna("Unknown")

rows = []
for _, r in df.iterrows():
    milestone = str(r["Milestone"]).strip()
    size = float(r["Estimate"])
    assignees = [a.strip() for a in str(r["Assignees"]).split(",") if a.strip()]
    if not assignees:
        continue
    share = round(size / len(assignees), 2)
    for a in assignees:
        rows.append({"Milestone": milestone, "Assignee": a, "Points": share})

points_df = pd.DataFrame(rows)
summary = (
    points_df.groupby(["Milestone", "Assignee"])["Points"]
    .sum()
    .reset_index()
)

milestone_totals = summary.groupby("Milestone")["Points"].sum().reset_index(name="TotalPoints")
summary = summary.merge(milestone_totals, on="Milestone")
summary["Percent"] = (summary["Points"] / summary["TotalPoints"] * 100)

pivot = summary.pivot_table(
    index="Milestone",
    columns="Assignee",
    values="Points",
    aggfunc="sum",
    fill_value=0
)

pivot = pivot.round(2)  
pivot["Total Points"] = milestone_totals.set_index("Milestone")["TotalPoints"].round(2)

percentages = (
    summary.pivot_table(
        index="Milestone",
        columns="Assignee",
        values="Percent",
        aggfunc="sum",
        fill_value=0
    )
    .add_suffix(" (%)")
    .round(1)
)

final = pd.concat([pivot, percentages], axis=1)

start_row = ws.max_row + 3  
start_col = 3  

header_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid") 
light_blue_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")  
medium_blue_fill = PatternFill(start_color="FFB4C6E7", end_color="FFB4C6E7", fill_type="solid") 

for col_idx, col_name in enumerate(final.reset_index().columns, start=start_col):
    cell = ws.cell(row=start_row, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = header_fill

for row_idx, row_data in enumerate(final.reset_index().itertuples(index=False), start=start_row + 1):
    for col_idx, val in enumerate(row_data, start=start_col):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if (col_idx - start_col) % 2 == 0:
            cell.fill = light_blue_fill
        else:
            cell.fill = medium_blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(val, (int, float)):
            cell.number_format = "0.00"  
            cell.font = Font(bold=True, color="000000")
        else:
            cell.font = Font(color="000000")

for column_cells in ws.columns:
    column_letter = get_column_letter(column_cells[0].column)
    max_length = 0
    for cell in column_cells:
        try:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

wb.save(OUTPUT_FILE)
print(f"Created '{OUTPUT_FILE}'")


